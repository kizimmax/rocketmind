import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FOLDER = os.path.dirname(os.path.abspath(__file__))

PAGES = [
    {
        "file": "rocketmind-2026-03-25 (9).json",
        "sheet": "Практикум бизнес-дизайн",
        "blocks": [
            ("Герой", ["ПРАКТИКУМ ПО БИЗНЕС-ДИЗАЙНУ ДЛЯ КОМАНД", "Освойте ключевые навыки стратегического развития бизнеса — от поиска бизнес-модели до проектирования платформ и экосистем."]),
            ("Доверие (цифры)", ["1000+ руководителей прошли обучение", "130+ часов консультаций проведено"]),
            ("Что такое практикум", ["Изучите фреймворки для бизнес-дизайна.\nПроанализируете текущую бизнес-модель вашей компании.\nСформулируете ценностное предложение для ваших клиентов.\nСоберёте канвасы платформы и экосистемы."]),
            ("Как проходит обучение", ["Доступ к образовательной платформе и шаблонам в Холсте.\nИнструкции и видеоуроки по каждому модулю.\nКейсы от крупных компаний и разборы реальных задач.\nОбратная связь от экспертов по вашим материалам."]),
            ("Кому подходит", ["ПРЕДПРИНИМАТЕЛЯМ — кто хочет переосмыслить бизнес-модель или запустить новый продукт.", "СТРАТЕГАМ И ИНВЕСТОРАМ — кто работает с портфелем компаний и хочет оценивать их потенциал."]),
            ("Что вы получите", ["Понимание полного процесса проектирования бизнес-модели.\nИнструменты бизнес-дизайна для применения в реальных проектах."]),
            ("Из чего состоит", ["Онлайн-программа продолжительностью 2 месяца.\nПроектные доски в Miro для работы над своим кейсом."]),
            ("Программа курса", ["МОДУЛЬ 1 — Введение в бизнес-дизайн. Основные канвасы и фреймворки.", "МОДУЛЬ 4 — Целевая аудитория. Ценностные предложения. Линейная бизнес-модель. Платформы.", "МОДУЛЬ 7 — Экосистемы.", "МОДУЛЬ 10 — ИИ в бизнес-дизайне."]),
            ("Эксперты", ["АЛЕКСЕЙ ЕРЁМИН — CEO Rocketmind, бизнес-дизайнер и стратег. 15+ лет в IT. Опыт с Росатом, Сбер, Газпромнефть, Норникель.", "АЛЕКСАНДР ПАВЛОВИЧ — продуктовый и сервисный дизайн, 15+ лет опыта."]),
            ("CTA", ["ГОТОВЫ ТРАНСФОРМИРОВАТЬ СВОЙ БИЗНЕС?", "Кнопка: ОСТАВИТЬ ЗАЯВКУ"]),
        ]
    },
    {
        "file": "rocketmind-2026-03-25 (10).json",
        "sheet": "Быстрый старт курс",
        "blocks": [
            ("Герой", ["БИЗНЕС-ДИЗАЙН: БЫСТРЫЙ СТАРТ", "Как мыслить стратегически в мире перемен?", "Цена: 10 000 ₽ (зачёркнутая цена: 15 000 ₽)"]),
            ("Формат", ["Минимум времени — максимум смысла\n2 ЧАСА / 6 ВИДЕО"]),
            ("Что узнаете", ["Что такое бизнес-дизайн и как мыслят успешные компании.\nПримеры компаний, которые переосмыслили свою бизнес-модель.\nМетоды пересборки бизнес-модели в условиях неопределённости.\nНавыки стратегического мышления для карьеры."]),
            ("Кому подойдёт", ["УПРАВЛЕНЦАМ — кто хочет быстро освоить инструменты стратегического мышления."]),
            ("Формат и длительность", ["6 видеоуроков по 10 минут каждый."]),
            ("Программа курса", ["УРОК 1 — Введение в бизнес-дизайн.", "УРОК 4 — Практика применения инструментов."]),
            ("Почему стоит пройти", ["Методология применялась в 100+ проектах для Сбер, Росатом, Газпромнефть.\nЗа 2 часа узнаете как запускают экосистемы, внедряют ИИ, переосмысливают бизнес-модели."]),
            ("Преподаватель", ["АЛЕКСЕЙ ЕРЁМИН — бизнес-дизайнер и стратег, эксперт по цифровым платформам и экосистемам."]),
            ("FAQ", ["Обучение проходит через платформу Skillspace.\nДоступ предоставляется сразу после оплаты.\nОплата картой."]),
            ("CTA", ["СДЕЛАЙТЕ ПЕРВЫЙ ШАГ К СТРАТЕГИЧЕСКОМУ МЫШЛЕНИЮ НОВОГО УРОВНЯ", "Цена: 10 000 ₽", "Кнопка: ОСТАВИТЬ ЗАЯВКУ"]),
        ]
    },
]


def style_header(ws, cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1a1a2e")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_block_name(cell):
    cell.font = Font(bold=True, size=10, color="1a1a2e")
    cell.fill = PatternFill("solid", fgColor="e8f0fe")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_content(cell):
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def thin_border():
    side = Side(style="thin", color="cccccc")
    return Border(left=side, right=side, top=side, bottom=side)


def make_sheet(wb, page):
    ws = wb.create_sheet(title=page["sheet"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    ws.append(["Название блока", "Контент"])
    for cell in ws[1]:
        style_header(ws, cell)
    ws.row_dimensions[1].height = 25

    row = 2
    for block_name, content_items in page["blocks"]:
        content_text = "\n\n".join(content_items)
        ws.cell(row=row, column=1, value=block_name)
        ws.cell(row=row, column=2, value=content_text)
        style_block_name(ws.cell(row=row, column=1))
        style_content(ws.cell(row=row, column=2))
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = thin_border()
        line_count = content_text.count("\n") + content_text.count("•") + 2
        ws.row_dimensions[row].height = max(30, min(line_count * 14, 200))
        row += 1

    return ws


wb = Workbook()
wb.remove(wb.active)

for page in PAGES:
    make_sheet(wb, page)

out_path = os.path.join(FOLDER, "rocketmind_academy_content.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
