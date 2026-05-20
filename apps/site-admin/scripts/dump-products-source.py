#!/usr/bin/env python3
"""
Reads docs/product_new/{xlsx,csv} and emits a normalized JSON:
[{ category, slug, sourceTitle, blocks: {ru_label: text} }, ...]

Output: apps/site-admin/scripts/.products-source.json
"""
import csv
import json
import os
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "docs" / "product_new"
OUT = Path(__file__).resolve().parent / ".products-source.json"

# Sheet name → (category, slug)
SHEET_TO_SLUG = {
    "ЭКОСИСТЕМНАЯ СТРАТЕГИЯ":            ("consulting", "ecosystem-strategy"),
    "СТРАТЕГИЧЕСКИЕ И ДИЗАЙН-СЕССИИ":    ("consulting", "strategy-sessions"),
    "СТАТУС РЕЗИДЕНТА «СКОЛКОВО»":       ("consulting", "skolkovo"),
    "УМНАЯ АНАЛИТИКА":                    ("consulting", "smart-analytics"),
    "ДИАГНОСТИКА ГОТОВНОСТИ КОМАНДЫ":    ("consulting", "team-readiness"),
    "ЗАПУСК ЦИФРОВОЙ ПЛАТФОРМЫ":         ("consulting", "digital-platform"),
}

# Standalone CSVs
CSV_AI = ("AИ-продукты - ИИзация вашего бизнеса.csv", "ai-products", "iizaciya-vashego-biznesa", "ИИзация вашего бизнеса")
CSV_ACADEMY = ("Онлайн-школа, продукты - ПРАКТИКУМ ПО БИЗНЕС-ДИЗАЙНУ.csv", "academy", "business-design-teams", "ПРАКТИКУМ ПО БИЗНЕС-ДИЗАЙНУ")


def normalize_text(t):
    if t is None:
        return ""
    if not isinstance(t, str):
        t = str(t)
    # Normalize NBSP and stray leading-space-of-continuation-lines
    return t.replace("\xa0", " ").strip()


def collect_blocks_from_rows(rows):
    """
    Rows: list[ (label, text) ]. Empty label = continuation of previous block.
    Returns dict {label: combined_text}.
    """
    blocks = {}
    current_label = None
    for raw_label, raw_text in rows:
        label = normalize_text(raw_label)
        text = normalize_text(raw_text)
        if label:
            current_label = label
            if current_label not in blocks:
                blocks[current_label] = text
            else:
                # Same label appears twice — append
                blocks[current_label] = (blocks[current_label] + "\n" + text).strip()
        elif current_label and text:
            # Continuation row
            blocks[current_label] = (blocks[current_label] + "\n" + text).strip()
    return blocks


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip()
        mapping = SHEET_TO_SLUG.get(key)
        if not mapping:
            print(f"  ⚠ sheet skipped (no slug mapping): {sheet_name!r}")
            continue
        category, slug = mapping
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Skip header row [0]
        block_rows = [(r[0], r[1]) if len(r) >= 2 else (r[0] if r else None, None) for r in rows[1:]]
        blocks = collect_blocks_from_rows(block_rows)
        out.append({
            "category": category,
            "slug": slug,
            "sourceTitle": key,
            "source": str(path.relative_to(ROOT)),
            "blocks": blocks,
        })
    return out


def read_csv(path, category, slug, source_title):
    with open(path, encoding="utf-8") as f:
        rows = list(csv.reader(f))
    # Skip header row
    block_rows = [(r[0] if len(r) >= 1 else None, r[1] if len(r) >= 2 else None) for r in rows[1:]]
    blocks = collect_blocks_from_rows(block_rows)
    return {
        "category": category,
        "slug": slug,
        "sourceTitle": source_title,
        "source": str(path.relative_to(ROOT)),
        "blocks": blocks,
    }


def main():
    items = []
    xlsx_path = SRC / "Консталтинг и стратегия, продукты.xlsx"
    print(f"Reading {xlsx_path.name}…")
    items.extend(read_xlsx(xlsx_path))
    print(f"Reading {CSV_AI[0]}…")
    items.append(read_csv(SRC / CSV_AI[0], CSV_AI[1], CSV_AI[2], CSV_AI[3]))
    print(f"Reading {CSV_ACADEMY[0]}…")
    items.append(read_csv(SRC / CSV_ACADEMY[0], CSV_ACADEMY[1], CSV_ACADEMY[2], CSV_ACADEMY[3]))

    OUT.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nWrote {len(items)} products to {OUT.relative_to(ROOT)}")
    for it in items:
        print(f"  {it['category']:>13}/{it['slug']:<32}  blocks: {len(it['blocks'])}  ({list(it['blocks'].keys())})")


if __name__ == "__main__":
    main()
